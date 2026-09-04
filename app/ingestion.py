import re
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from zipfile import BadZipFile

from langchain_core.documents import Document
from langchain_text_splitters import RecursiveCharacterTextSplitter
from pypdf import PdfReader
from pypdf.errors import PdfReadError

ALLOWED_EXTENSIONS = {".txt", ".md", ".pdf", ".docx", ".xlsx"}
ALLOWED_CONTENT_TYPES = {
    ".txt": {"text/plain", "application/octet-stream"},
    ".md": {"text/markdown", "text/plain", "application/octet-stream"},
    ".pdf": {"application/pdf", "application/octet-stream"},
    ".docx": {
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        "application/octet-stream",
    },
    ".xlsx": {
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/octet-stream",
    },
}


@dataclass(frozen=True)
class DocumentPage:
    page_number: int
    text: str


def validate_upload(
    filename: str,
    size: int,
    content_type: str | None,
    max_bytes: int,
) -> str:
    if not filename or filename != Path(filename).name or "\x00" in filename:
        raise ValueError("Filename must be a simple file name without path components.")
    suffix = Path(filename).suffix.lower()
    if suffix not in ALLOWED_EXTENSIONS:
        raise ValueError("Unsupported file type. Upload a .txt, .md, or .pdf file.")
    if size <= 0:
        raise ValueError("The uploaded document is empty.")
    if size > max_bytes:
        raise ValueError(f"The uploaded document exceeds the {max_bytes} byte limit.")
    if content_type and content_type not in ALLOWED_CONTENT_TYPES[suffix]:
        raise ValueError("The file content type does not match its extension.")
    return filename


def clean_text(text: str) -> str:
    text = text.replace("\r\n", "\n").replace("\r", "\n").replace("\u00a0", " ")
    lines = [re.sub(r"[ \t]+", " ", line).strip() for line in text.split("\n")]
    return "\n".join(line for line in lines if line).strip()


def extract_document_pages(filename: str, content: bytes) -> list[DocumentPage]:
    suffix = Path(filename).suffix.lower()
    if suffix not in ALLOWED_EXTENSIONS:
        raise ValueError("Unsupported file type. Upload a .txt, .md, or .pdf file.")
    try:
        if suffix in {".txt", ".md"}:
            pages = [DocumentPage(1, clean_text(content.decode("utf-8")))]
        elif suffix == ".pdf":
            reader = PdfReader(BytesIO(content))
            pages = [
                DocumentPage(index, clean_text(page.extract_text() or ""))
                for index, page in enumerate(reader.pages, start=1)
            ]
        elif suffix == ".docx":
            from docx import Document as WordDocument

            document = WordDocument(BytesIO(content))
            pages = [DocumentPage(1, clean_text("\n".join(paragraph.text for paragraph in document.paragraphs)))]
        elif suffix == ".xlsx":
            from openpyxl import load_workbook

            workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
            pages = []
            for index, worksheet in enumerate(workbook.worksheets, start=1):
                rows = ("\t".join("" if value is None else str(value) for value in row) for row in worksheet.iter_rows(values_only=True))
                pages.append(DocumentPage(index, clean_text("\n".join(rows))))
            workbook.close()
    except (BadZipFile, PdfReadError, UnicodeDecodeError, ValueError) as error:
        raise ValueError("The document could not be read. Upload a valid supported document.") from error
    pages = [page for page in pages if page.text]
    if not pages:
        raise ValueError("The document contains no extractable text.")
    return pages


def chunk_document(
    filename: str,
    pages: list[DocumentPage],
    splitter: RecursiveCharacterTextSplitter,
) -> list[Document]:
    chunks: list[Document] = []
    chunk_index = 0
    for page in pages:
        page_chunks = splitter.create_documents(
            [page.text], metadatas=[{"source": filename, "page": page.page_number}]
        )
        for document in page_chunks:
            document.metadata["chunk"] = chunk_index
            chunks.append(document)
            chunk_index += 1
    return chunks
